import openpyxl
import random
import string

import pandas as pd


class DataGenerator:
    """
    Generates an Excel file with two sheets of dummy data for testing.
    """

    def __init__(self, filename="data.xlsx"):
        self.filename = filename

    def generate_random_string(self, length=10):
        """
        Generate a random string of specified length.

        Parameters:
            length (int): Length of the generated string.

        Returns:
            str: Randomly generated string.
        """
        letters = string.ascii_letters
        return ''.join(random.choice(letters) for i in range(length))

    def generate_random_code(self):
        """
        Generate a random code in the format NNNN-DDD.

        Returns:
            str: Randomly generated code.
        """
        numbers = ''.join(random.choice('01') for _ in range(4))  # 4 digits (0 or 1)
        letters = ''.join(random.choice('ABC') for _ in range(3))  # 3 letters (A, B, C)
        return f"{numbers}-{letters}"

    # ! Using openpyxl:
    def create_excel_file(self):
        """
        Create an Excel file with two sheets, 'operations' and 'materials',
        each populated with headers and dummy data.
        """
        workbook = openpyxl.Workbook()

        # Create the 'operations' sheet
        operations_sheet = workbook.active
        operations_sheet.title = "operations"

        # Add headers to 'operations' sheet
        for col in range(1, 11):
            operations_sheet.cell(row=1, column=col, value=f'col{col}')

        # Fill 'operations' sheet with dummy data
        for row in range(2, 52):  # 50 rows of data
            for col in range(1, 11):
                if col == 5:
                    operations_sheet.cell(row=row, column=col, value=self.generate_random_code())
                else:
                    operations_sheet.cell(row=row, column=col, value=self.generate_random_string())

        # Create the 'materials' sheet
        materials_sheet = workbook.create_sheet(title="materials")

        # Add headers to 'materials' sheet
        for col in range(1, 11):
            materials_sheet.cell(row=1, column=col, value=f'col{col}')

        # Fill 'materials' sheet with dummy data
        # ? Should we add a progress indicator here?
        for row in range(2, 10000):  # 1 million rows of data
            for col in range(1, 11):
                if col == 10:
                    materials_sheet.cell(row=row, column=col, value=self.generate_random_code())
                else:
                    materials_sheet.cell(row=row, column=col, value=self.generate_random_string())

        # Save the file
        workbook.save(self.filename)




    # ! Using pandas:
    def create_excel_file2(self):
        """
        Create an Excel file with two sheets, 'operations' and 'materials',
        each populated with headers and dummy data. Uses pandas for faster performance with large datasets.
        """
        # Define the number of rows for each sheet
        # Define row counts for sheets
        operations_rows = 50
        materials_rows = 1_000_000

        # Generate 'operations' data
        operations_data = {
            f"col{col}": [
                self.generate_random_code() if col == 5 else self.generate_random_string()
                for _ in range(operations_rows)
            ]
            for col in range(1, 11)
        }
        operations_df = pd.DataFrame(operations_data)

        # Initialize Excel writer
        with pd.ExcelWriter(self.filename, engine="xlsxwriter") as writer:
            operations_df.to_excel(writer, sheet_name="operations", index=False)

            # For 'materials', generate and write in chunks
            chunk_size = 100_000  # Adjust chunk size if needed for memory usage
            for start_row in range(0, materials_rows, chunk_size):
                end_row = min(start_row + chunk_size, materials_rows)
                materials_data = {
                    f"col{col}": [
                        self.generate_random_code() if col == 10 else self.generate_random_string()
                        for _ in range(start_row, end_row)
                    ]
                    for col in range(1, 11)
                }
                materials_df = pd.DataFrame(materials_data)

                # Write each chunk to the same sheet, starting at appropriate row
                materials_df.to_excel(writer, sheet_name="materials", index=False, startrow=start_row)
