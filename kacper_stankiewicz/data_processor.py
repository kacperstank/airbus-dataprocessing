import pandas as pd
import openpyxl
import time


class DataProcessor:
    """
    Processes Excel data by extracting unique values and counting occurrences
    using openpyxl, pandas or xlsxio.
    """

    def __init__(self, filename="data.xlsx"):
        self.filename = filename

    def extract_unique_values(self):
        """
        Extract unique values from column 5 in the 'operations' sheet.

        Returns:
            set: Unique values extracted from column 5.
        """
        try:
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(self.filename, data_only=True)
            operations_sheet = workbook["operations"]

            # Use a set to store unique values from column 5 (E)
            unique_values = set()

            for row in operations_sheet.iter_rows(
                    min_row=2, max_row=51, min_col=5, max_col=5, values_only=True
            ):
                cell_value = row[0]
                # Add non-empty values to the set
                if cell_value:
                    unique_values.add(cell_value)

            workbook.close()

            return unique_values

        except Exception as e:
            print(f"Error in extract_unique_values: {e}")
            return set()

    def count_occurrences(self, unique_values):
        """
        Count occurrences of each unique value in column 10 of the 'materials' sheet.

        Parameters:
            unique_values (set): Set of unique values to count.

        Returns:
            tuple: Dictionary of counts and the time taken for processing.
        """
        start_time = time.time()  # Start timing

        try:
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(self.filename, data_only=True)
            materials_sheet = workbook["materials"]

            # Initialize a dictionary to store counts for each unique value
            counts = {value: 0 for value in unique_values}

            for row in materials_sheet.iter_rows(
                    min_row=2, min_col=10, max_col=10, values_only=True
            ):
                cell_value = row[0]

                # Increment the count if the cell value is in unique_values
                if cell_value in counts:
                    counts[cell_value] += 1

            workbook.close()

            time_taken = time.time() - start_time  # Calculate time taken
            return counts, time_taken  # Return results and time taken

        except Exception as e:
            print(f"Error in count_occurrences: {e}")
            return {}

    def process_with_pandas(self):
        """
        Process data using pandas to improve performance with large datasets.

        Returns:
            tuple: Dictionary of counts and the time taken for processing.
        """
        start_time = time.time()  # Start timing

        try:
            # Extract unique values from 'operations' sheet column 5
            operations_df = pd.read_excel(self.filename, sheet_name="operations", usecols=[4])
            unique_values = operations_df.iloc[:, 0].dropna().unique()

            # Count occurrences in 'materials' sheet column 10
            materials_df = pd.read_excel(self.filename, sheet_name="materials", usecols=[9])
            counts = {value: (materials_df.iloc[:, 0] == value).sum() for value in unique_values}

            time_taken = time.time() - start_time  # Calculate time taken
            return counts, time_taken  # Return results and time taken

        except Exception as e:
            print(f"Error in process_with_pandas: {e}")
            return {}

    def process_with_xlsxio(self):
        """
        Placeholder for xlsxio processing (requires C bindings).

        Returns:
            dict: Placeholder empty dictionary for future xlsxio processing.
        """
        print("xlsxio processing is not implemented due to lack of Python bindings.")
        return {}