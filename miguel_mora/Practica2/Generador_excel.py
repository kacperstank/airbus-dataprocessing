import openpyxl
import random
import string
class Generador_excel:
    # Función para generar un string aleatorio
    def random_string(length=10):
        letters = string.ascii_letters
        return ''.join(random.choice(letters) for i in range(length))

    # Función para generar un formato NNNN-DDD
    def random_code(self):
        numbers = ''.join(random.choice('01') for _ in range(4))  # 4 dígitos (0 o 1)
        letters = ''.join(random.choice('ABC') for _ in range(3))  # 3 letras (A, B, C)
        return f"{numbers}-{letters}"

    # Crear un nuevo libro de trabajo
    wb = openpyxl.Workbook()

    # Crear la hoja "operaciones"
    ws1 = wb.active
    ws1.title = "operaciones"
    filas = 100000
    print(f"[!] Creando {filas} de filas en 'datos.xlsx'... Espera un poco")


    # Crear las columnas
    for col in range(1, 11):
        ws1.cell(row=1, column=col, value=f'col{col}')

    # Rellenar con datos aleatorios
    for row in range(2, 52):  # 50 filas de datos
        for col in range(1, 11):
            if col == 5:
                ws1.cell(row=row, column=col, value=random_code())  # Columna 5 con formato NNNN-DDD
            else:
                ws1.cell(row=row, column=col, value=random_string())

    # Crear la hoja "materiales"
    ws2 = wb.create_sheet(title="materiales")

    # Crear las columnas
    for col in range(1, 11):
        ws2.cell(row=1, column=col, value=f'col{col}')

    # Rellenar con datos aleatorios
    for row in range(2, filas):  # 1 millón de filas de datos
        for col in range(1, 11):
            if col == 10:
                ws2.cell(row=row, column=col, value=random_code())  # Columna 10 con formato NNNN-DDD
            else:
                ws2.cell(row=row, column=col, value=random_string())


    # Guardar el archivo
    wb.save("datos.xlsx")

    print("[OK] Archivo 'datos.xlsx' creado con éxito.")
